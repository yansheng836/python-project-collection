#抓取京东商品详情页数据
import  requests
from bs4 import  BeautifulSoup
import openpyxl
import  time
import  re
import json
#搜索商品列表页的每个商品的链接
def make_a_link(keyword,page):
    url = 'https://search.jd.com/Search?keyword=' + keyword + '&enc=utf-8&page=' + str(page * 2 - 1)
    res = requests.get(url)
    res.raise_for_status()
    res.encoding = res.apparent_encoding  # 转码 为防止出现乱码
    print('正在爬取第' + str(page) + '页:' + url)
    html = res.text
    soup = BeautifulSoup(html, 'lxml')
    links = soup.find_all('li', class_='gl-item')  # 所有商品的li
    return (link for link in links)
#详情页
def detail_link(purl):
    try:
        r = requests.get(purl)
        r.raise_for_status
        r.encoding = 'gbk'
        return r.text
    except:
        print('此页无法链接！！！')
        return ''
#商品的名字和价格
def get_name_price(uid,row,sheet):
    content = detail_link('https://c.3.cn/recommend?&methods=accessories&sku=' + uid + '&cat=9987%2C653%2C655')
    try:
        jd=json.loads(content)
        sheet.cell(row=row,column=2)._value=jd['accessories']['data']['wName']
        sheet.cell(row=row,column=3)._value=jd['accessories']['data']['wMaprice']
        print(jd['accessories']['data']['wName'])
    except:
        return ""
#店铺
def get_shop(uid,row,sheet):
    content = detail_link('https://chat1.jd.com/api/checkChat?pid=' + uid + '&returnCharset=utf-8')
    try:
        jd = json.loads(content.lstrip('null(').rstrip(');'))
        try:
            sheet.cell(row=row, column=4)._value = jd['seller']
        except:
            return ''
    except:
        ''
#商品的评论
def get_comments(uid,row,sheet):
    content = detail_link('https://club.jd.com/comment/productCommentSummaries.action?referenceIds=' + uid)
    jd=json.loads(content)
    sheet.cell(row=row,column=6)._value=jd['CommentsCount'][0]['CommentCountStr'] #总评
    sheet.cell(row=row,column=7)._value=jd['CommentsCount'][0]['GoodCountStr'] #好评
    sheet.cell(row=row,column=8)._value=jd['CommentsCount'][0]['GoodRate'] #好评率
def main():
    wb =openpyxl.Workbook()
    sheet=wb.active
    sheet.title = "京东抓取商品数据.xlsx"
    sheet.cell(row=1,column=1)._value='商品ID'
    sheet.cell(row=1,column=2)._value='商品名称'
    sheet.cell(row=1,column=3)._value= '价格'
    sheet.cell(row=1,column=4)._value='店铺'
    sheet.cell(row=1,column=5)._value= '链接'
    sheet.cell(row=1,column=6)._value='评论数'
    sheet.cell(row=1,column=7)._value='好评数'
    sheet.cell(row=1,column=8)._value= '评论率'
    row=2
    keyword=input("请输入要抓取的商品：")
    pages=input("要抓取的页数：")   #str类型
    starttime=time.time()
    pages=int(pages)
    for page in range(1,pages+1):
        for link in make_a_link(keyword, page):
            uid=link['data-sku']
            # uid = re.match(r'.+?(\d+).+', purl).group(1)  # 商品的ID
            sheet.cell(row=row, column=1)._value =uid
            purl=link.find('div', class_='p-name p-name-type-2').a['href']
            if 'http' not in purl:
                sheet.cell(row=row, column=5)._value = 'http:' + purl
            else:
                sheet.cell(row=row, column=5)._value = purl
            get_name_price(uid, row, sheet)
            get_shop(uid, row, sheet)
            get_comments(uid, row, sheet)
            row+=1
    wb.save('京东抓取'+keyword+'数据.xlsx')
    print('耗时{}秒。'.format(time.time() - starttime))  # 爬取所需时间
if __name__ == '__main__':
    main()